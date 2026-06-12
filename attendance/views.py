from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages
import random, string, base64, math
from .models import (Student, AdminProfile, AdminOTP, DailyCode,
                     OTP, Attendance, COLLEGE_LAT, COLLEGE_LNG, ALLOWED_RADIUS_METERS)

# ── Haversine: distance between two GPS points in meters ──────────────────────
def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

# ══════════════════════════════════════════════════════════════════════════════
#  STUDENT FLOW
# ══════════════════════════════════════════════════════════════════════════════

def index(request):
    return render(request, 'attendance/index.html')

def verify_roll(request):
    if request.method == 'POST':
        roll = request.POST.get('roll_number', '').strip().upper()
        try:
            student = Student.objects.get(roll_number=roll)
            otp_code = ''.join(random.choices(string.digits, k=6))
            OTP.objects.create(student=student, otp_code=otp_code)
            request.session['student_id'] = student.id
            request.session['otp_code']   = otp_code   # demo mode
            return redirect('verify_otp')
        except Student.DoesNotExist:
            return render(request, 'attendance/index.html',
                          {'error': 'Roll number not found. Contact admin.'})
    return redirect('index')

def verify_otp(request):
    student_id = request.session.get('student_id')
    otp_demo   = request.session.get('otp_code')
    if not student_id:
        return redirect('index')
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        entered = request.POST.get('otp', '').strip()
        try:
            otp_obj = OTP.objects.filter(student=student, is_used=False).latest('created_at')
            if otp_obj.is_valid() and otp_obj.otp_code == entered:
                otp_obj.is_used = True
                otp_obj.save()
                return redirect('mark_attendance')
            else:
                return render(request, 'attendance/otp.html',
                              {'student': student,
                               'error': 'Invalid or expired OTP. Try again.',
                               'otp_demo': otp_demo})
        except OTP.DoesNotExist:
            return render(request, 'attendance/otp.html',
                          {'student': student,
                           'error': 'No OTP found. Go back and retry.',
                           'otp_demo': otp_demo})
    return render(request, 'attendance/otp.html',
                  {'student': student, 'otp_demo': otp_demo})

def mark_attendance(request):
    student_id = request.session.get('student_id')
    if not student_id:
        return redirect('index')
    student = get_object_or_404(Student, id=student_id)
    today   = timezone.now().date()

    if Attendance.objects.filter(student=student, date=today).exists():
        return render(request, 'attendance/already_marked.html', {'student': student})

    daily_code = DailyCode.get_or_create_today()

    if request.method == 'POST':
        entered_code = request.POST.get('daily_code', '').strip()
        lat_str      = request.POST.get('latitude', '')
        lng_str      = request.POST.get('longitude', '')
        photo_data   = request.POST.get('photo_data', '')

        # ── Daily code check
        if entered_code != daily_code.code:
            return render(request, 'attendance/mark.html',
                          {'student': student, 'daily_code': daily_code,
                           'error': 'Wrong daily code! Ask your teacher.',
                           'college_lat': COLLEGE_LAT, 'college_lng': COLLEGE_LNG})

        # ── GPS check
        status     = 'Absent'
        distance_m = None
        lat = lng = None

        if lat_str and lng_str:
            try:
                lat = float(lat_str)
                lng = float(lng_str)
                distance_m = haversine(lat, lng, COLLEGE_LAT, COLLEGE_LNG)
                status = 'Present' if distance_m <= ALLOWED_RADIUS_METERS else 'Absent'
            except ValueError:
                pass

        att = Attendance.objects.create(
            student    = student,
            date       = today,
            daily_code = daily_code,
            latitude   = lat,
            longitude  = lng,
            distance_m = distance_m,
            status     = status,
        )

        # ── Save photo
        if photo_data and photo_data.startswith('data:image'):
            from django.core.files.base import ContentFile
            fmt, imgstr = photo_data.split(';base64,')
            ext = fmt.split('/')[-1]
            photo_file = ContentFile(
                base64.b64decode(imgstr),
                name=f"att_{student.roll_number}_{today}.{ext}"
            )
            att.photo.save(photo_file.name, photo_file, save=True)

        request.session.flush()
        return redirect('success' if status == 'Present' else 'marked_absent')

    return render(request, 'attendance/mark.html', {
        'student': student, 'daily_code': daily_code,
        'college_lat': COLLEGE_LAT, 'college_lng': COLLEGE_LNG,
        'allowed_radius': ALLOWED_RADIUS_METERS,
    })

def success(request):
    return render(request, 'attendance/success.html', {'status': 'Present'})

def marked_absent(request):
    return render(request, 'attendance/success.html', {'status': 'Absent'})

# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN FLOW
# ══════════════════════════════════════════════════════════════════════════════

def admin_login(request):
    if request.session.get('admin_logged_in'):
        return redirect('admin_dashboard')
    if request.method == 'POST':
        # Generate OTP and store in session (demo: show on screen)
        otp_code = ''.join(random.choices(string.digits, k=6))
        AdminOTP.objects.create(otp_code=otp_code)
        request.session['admin_otp'] = otp_code
        return redirect('admin_verify_otp')
    return render(request, 'attendance/admin_login.html')

def admin_verify_otp(request):
    otp_demo = request.session.get('admin_otp')
    admin    = AdminProfile.objects.first()
    if request.method == 'POST':
        entered = request.POST.get('otp', '').strip()
        try:
            otp_obj = AdminOTP.objects.filter(is_used=False).latest('created_at')
            if otp_obj.is_valid() and otp_obj.otp_code == entered:
                otp_obj.is_used = True
                otp_obj.save()
                request.session['admin_logged_in'] = True
                return redirect('admin_dashboard')
            else:
                return render(request, 'attendance/admin_otp.html',
                              {'error': 'Invalid or expired OTP.', 'otp_demo': otp_demo, 'admin': admin})
        except AdminOTP.DoesNotExist:
            return render(request, 'attendance/admin_otp.html',
                          {'error': 'No OTP found.', 'otp_demo': otp_demo, 'admin': admin})
    return render(request, 'attendance/admin_otp.html', {'otp_demo': otp_demo, 'admin': admin})

def admin_logout(request):
    request.session.flush()
    return redirect('admin_login')

def admin_dashboard(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    today       = timezone.now().date()
    daily_code  = DailyCode.get_or_create_today()
    admin       = AdminProfile.objects.first()
    students    = Student.objects.all()
    today_att   = Attendance.objects.filter(date=today).select_related('student')
    present_ids = set(a.student_id for a in today_att if a.status == 'Present')
    absent_ids  = set(a.student_id for a in today_att if a.status == 'Absent')
    not_marked  = students.exclude(id__in=present_ids).exclude(id__in=absent_ids)

    return render(request, 'attendance/admin_dashboard.html', {
        'daily_code'    : daily_code,
        'admin'         : admin,
        'today_att'     : today_att,
        'not_marked'    : not_marked,
        'total'         : students.count(),
        'present_count' : len(present_ids),
        'absent_count'  : len(absent_ids) + not_marked.count(),
        'today'         : today,
        'college_lat'   : COLLEGE_LAT,
        'college_lng'   : COLLEGE_LNG,
        'allowed_radius': ALLOWED_RADIUS_METERS,
    })
# ================================
# EXCEL EXPORT
# ================================

def export_excel(request):
    if not request.session.get('admin_logged_in'):
        return redirect('admin_login')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    today = timezone.now().date()
    admin = AdminProfile.objects.first()
    students = Student.objects.all().order_by('roll_number')
    today_att = {a.student_id: a for a in Attendance.objects.filter(date=today)}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {today}"

    ws.merge_cells('A1:G1')
    ws['A1'] = f"SmartAttend — {admin.section if admin else ''}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['#', 'Roll No', 'Name', 'Phone', 'Status', 'Time', 'Distance']
    ws.append([])
    ws.append(headers)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for idx, s in enumerate(students, start=1):
        att = today_att.get(s.id)
        if att:
            status = att.status
            time_str = att.marked_at.strftime('%H:%M:%S')
            dist = f"{att.distance_m:.0f}" if att.distance_m else "-"
        else:
            status = "Absent"
            time_str = "-"
            dist = "-"

        row = [idx, s.roll_number, s.name, s.phone_number, status, time_str, dist]
        ws.append(row)
        status_cell = ws.cell(row=ws.max_row, column=5)
        if status == 'Present':
            status_cell.font = Font(color="10B981", bold=True)
        else:
            status_cell.font = Font(color="EF4444", bold=True)

    widths = [4, 14, 22, 16, 10, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.append([])
    ws.append(['', '', '', '', 'Total', '', students.count()])
    ws.append(['', '', '', '', 'Present', '', len(today_att)])
    ws.append(['', '', '', '', 'Absent', '', students.count() - len(today_att)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_{today}.xlsx"'
    wb.save(response)
    return response
